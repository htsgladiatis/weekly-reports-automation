"""
Парсер Excel выгрузок из Яндекс.Директ.

Читает файлы формата:
- 2026-06-08_22-11-46_e-17228851.xlsx
- 2026-06-08_22-13-28_dune-group.xlsx

Создает unified CSV для auto_weekly_report.py
"""

import os
import sys
import re
import csv
from typing import List, Dict, Any

try:
    import openpyxl
except ImportError:
    print("❌ Требуется openpyxl для чтения Excel файлов")
    print("   Установите: pip install openpyxl")
    sys.exit(1)


def extract_account_from_filename(filename: str) -> str:
    """
    Извлекает название аккаунта из имени файла.
    
    Examples:
        2026-06-08_22-11-46_e-17228851.xlsx → e-17228851
        2026-06-08_22-13-28_dune-group.xlsx → dune-group
    """
    match = re.search(r'_(e-\d+|dune-group|porg-[a-z0-9]+)\.xlsx$', filename)
    if match:
        return match.group(1)
    return "unknown"


def parse_xlsx_file(filepath: str, account: str) -> List[Dict[str, Any]]:
    """
    Парсит Excel файл выгрузки из Директа.
    
    Args:
        filepath: Путь к .xlsx файлу
        account: Название аккаунта (e-20010227, dune-group, etc.)
    
    Returns:
        Список кампаний с данными
    """
    campaigns = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        # Найти заголовки
        header_row = None
        headers = {}
        
        # Расширенный поиск, потому что структура XLSX может быть другой
        max_scan_rows = 120
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), 1):
            row_str = " | ".join([str(c).strip().lower() for c in row if c is not None])
            
            # Ищем строку, в которой потенциально есть колонка кампаний
            if 'камп' in row_str or 'campaign' in row_str:
                header_row = i
                
                # Маппинг заголовков
                for j, cell in enumerate(row):
                    if cell is None:
                        continue
                    
                    cell_lower = str(cell).strip().lower()
                    if not cell_lower:
                        continue
                    
                    # Campaign
                    if ('кампани' in cell_lower) or ('campaign' in cell_lower):
                        headers['campaign'] = j
                        continue
                    
                    # Impressions / Показы
                    if any(k in cell_lower for k in [
                        'показы', 'показов', 'показ', 'impressions', 'impression', 'views', 'view', 'impr'
                    ]):
                        headers['impressions'] = j
                        continue
                    
                    # Clicks / Клики
                    if any(k in cell_lower for k in [
                        'клики', 'клика', 'клик', 'clicks', 'click', 'ctr', 'chicks'
                    ]):
                        headers['clicks'] = j
                        continue
                    
                    # Spend / Расход
                    if any(k in cell_lower for k in [
                        'расход', 'стоимость', 'cost', 'spend', 'budget', 'сумма'
                    ]):
                        headers['spend'] = j
                        continue
                
                # Если нашли campaign — обычно этого достаточно, чтобы остановиться и проверить полноту
                break
        
        if not header_row:
            print(f"⚠️  Не найдены заголовки в {filepath} (кампания)")
            return campaigns
        
        required = ['campaign', 'impressions', 'clicks', 'spend']
        if not all(k in headers for k in required):
            # Логируем хотя бы то, что удалось извлечь из заголовков
            header_values = []
            for cell in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
                header_values = [str(c).strip() if c is not None else "" for c in cell]
                break
            print(
                f"⚠️  Неполные заголовки в {filepath}. "
                f"Найдено: {headers}. "
                f"Строка заголовков (row {header_row}): {header_values}"
            )
            return campaigns
        
        # Читаем данные
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[headers['campaign']]:
                continue
            
            campaign_name = str(row[headers['campaign']]).strip()
            
            # Пропускаем итоговые строки
            if any(word in campaign_name.lower() for word in ['итого', 'total', 'всего']):
                continue
            
            # Парсим числа
            try:
                impressions = int(row[headers['impressions']] or 0)
                clicks = int(row[headers['clicks']] or 0)
                
                # Расход может быть в разных форматах
                spend_raw = row[headers['spend']]
                if spend_raw is None:
                    spend = 0
                elif isinstance(spend_raw, (int, float)):
                    spend = float(spend_raw)
                else:
                    # Убираем пробелы и запятые
                    spend_str = str(spend_raw).replace(' ', '').replace(',', '.')
                    spend = float(spend_str) if spend_str else 0
                
                campaigns.append({
                    'account': account,
                    'campaign': campaign_name,
                    'impressions': impressions,
                    'clicks': clicks,
                    'spend': spend
                })
                
            except (ValueError, TypeError) as e:
                print(f"⚠️  Ошибка парсинга строки: {row} — {e}")
                continue
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Ошибка чтения {filepath}: {e}")
    
    return campaigns


def parse_directory(directory: str) -> List[Dict[str, Any]]:
    """
    Парсит все .xlsx файлы в директории.
    
    Args:
        directory: Путь к папке с выгрузками
    
    Returns:
        Объединенный список всех кампаний
    """
    all_campaigns = []
    
    if not os.path.exists(directory):
        print(f"❌ Директория не найдена: {directory}")
        return all_campaigns
    
    xlsx_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    
    if not xlsx_files:
        print(f"⚠️  Не найдено .xlsx файлов в {directory}")
        return all_campaigns
    
    print(f"📂 Найдено файлов: {len(xlsx_files)}")
    
    for filename in xlsx_files:
        filepath = os.path.join(directory, filename)
        account = extract_account_from_filename(filename)
        
        print(f"\n📊 Парсинг: {filename} (аккаунт: {account})")
        
        campaigns = parse_xlsx_file(filepath, account)
        
        if campaigns:
            print(f"   ✅ Загружено кампаний: {len(campaigns)}")
            all_campaigns.extend(campaigns)
        else:
            print(f"   ⚠️  Нет данных")
    
    return all_campaigns


def save_to_csv(campaigns: List[Dict[str, Any]], output_file: str):
    """Сохраняет данные в CSV файл."""
    if not campaigns:
        print("❌ Нет данных для сохранения")
        return False
    
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Account', 'Campaign', 'Impressions', 'Clicks', 'Spend']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for camp in campaigns:
                writer.writerow({
                    'Account': camp['account'],
                    'Campaign': camp['campaign'],
                    'Impressions': camp['impressions'],
                    'Clicks': camp['clicks'],
                    'Spend': f"{camp['spend']:.2f}"
                })
        
        print(f"\n✅ CSV сохранен: {output_file}")
        
        # Статистика
        total_spend = sum(c['spend'] for c in campaigns)
        total_clicks = sum(c['clicks'] for c in campaigns)
        total_impressions = sum(c['impressions'] for c in campaigns)
        
        print(f"\n📊 Итого:")
        print(f"   Кампаний: {len(campaigns)}")
        print(f"   Расход: {total_spend:,.2f} руб")
        print(f"   Показы: {total_impressions:,}")
        print(f"   Клики: {total_clicks:,}")
        
        if total_impressions > 0:
            ctr = (total_clicks / total_impressions * 100)
            print(f"   CTR: {ctr:.2f}%")
        
        if total_clicks > 0:
            cpc = total_spend / total_clicks
            print(f"   CPC: {cpc:.2f} руб")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка сохранения CSV: {e}")
        return False


def main():
    """CLI для парсинга Excel файлов."""
    print("📊 Парсер Excel выгрузок Яндекс.Директ")
    print("=" * 70)
    
    if len(sys.argv) < 3:
        print("\nИспользование:")
        print("  python parse_direct_xlsx.py <директория> <output.csv>")
        print("\nПример:")
        print("  python parse_direct_xlsx.py 1-7.06 direct_0106.csv")
        print("  python parse_direct_xlsx.py \"1-7.06\" direct_w5.csv")
        return 1
    
    directory = sys.argv[1]
    output_file = sys.argv[2]
    
    # Парсим все файлы
    campaigns = parse_directory(directory)
    
    if not campaigns:
        print("\n❌ Не удалось извлечь данные")
        return 1
    
    # Сохраняем в CSV
    if save_to_csv(campaigns, output_file):
        print(f"\n✅ Готово! Используйте файл:")
        print(f"   python auto_weekly_report.py 2026-06-01 2026-06-07 {output_file}")
        return 0
    else:
        return 1


if __name__ == "__main__":
    sys.exit(main())
