"""
Parse W9 data files for 29.06-05.07.2026
"""
import os
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not available")
    raise

try:
    import xlrd
except ImportError:
    xlrd = None
    print("xlrd not available - will try openpyxl for .xls")

PROJECT_ROOT = Path(".")
WEEK9_DIR = PROJECT_ROOT / "05.07"
DIRECT_DIR = WEEK9_DIR / "111"
WEBMASTER_DIR = WEEK9_DIR / "webmaster"

# 4 Direct accounts naming
ACCOUNT_MAP = {
    "e-20010227": "e-20010227",
    "e-17228851": "e-17228851",
    "dune-group": "dune-group",
    "porg-3uieikjn": "porg-3uieikjn",
}

results = {
    "period": "29.06.2026 - 05.07.2026",
    "direct_accounts": {},
    "total_impressions": 0,
    "total_clicks": 0,
    "total_spend": 0.0,
    "files": []
}

def detect_account(filename):
    for key in ACCOUNT_MAP:
        if key in filename:
            return key
    return None

def parse_xlsx(filepath):
    """Parse xlsx - return aggregated totals per campaign"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 5:
            return None

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and "Название кампании" in str(row[0]):
                header_idx = i
                break

        if header_idx is None:
            return None

        headers = list(rows[header_idx])

        # Build column map
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                col_map[str(h).strip()] = i

        # Required columns
        name_col = col_map.get("Название кампании")
        if name_col is None:
            return None

        spend_col = col_map.get("Расход, ₽")
        clicks_col = col_map.get("Клики")
        impressions_col = col_map.get("Показы")
        conversions_col = col_map.get("Конверсии")

        campaigns = {}
        for row in rows[header_idx + 1:]:
            if not row or row[name_col] is None:
                continue
            name = str(row[name_col]).strip()

            if name.lower() in ("итого", "total", ""):
                continue

            def num(v):
                if v in (None, "", "-"):
                    return 0
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(str(v).replace(",", ".").replace(" ", ""))
                except (ValueError, TypeError):
                    return 0

            spend = num(row[spend_col]) if spend_col is not None else 0
            clicks = num(row[clicks_col]) if clicks_col is not None else 0
            impressions = num(row[impressions_col]) if impressions_col is not None else 0
            conversions = num(row[conversions_col]) if conversions_col is not None else 0

            if name in campaigns:
                campaigns[name]["spend"] += spend
                campaigns[name]["clicks"] += clicks
                campaigns[name]["impressions"] += impressions
                campaigns[name]["conversions"] += conversions
            else:
                campaigns[name] = {
                    "spend": spend,
                    "clicks": clicks,
                    "impressions": impressions,
                    "conversions": conversions,
                }

        total_spend = sum(c["spend"] for c in campaigns.values())
        total_clicks = sum(c["clicks"] for c in campaigns.values())
        total_impressions = sum(c["impressions"] for c in campaigns.values())
        total_conversions = sum(c["conversions"] for c in campaigns.values())

        return {
            "campaigns": campaigns,
            "totals": {
                "spend": total_spend,
                "clicks": total_clicks,
                "impressions": total_impressions,
                "conversions": total_conversions,
            },
            "has_impressions_col": impressions_col is not None,
            "headers": headers,
        }
    except Exception as e:
        return {"error": str(e)}

def parse_seo_webmaster(filepath):
    """Parse Яндекс.Вебмастер/Метрика SEO data"""
    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        return {
            "rows": [
                {f"col_{i}": str(v) if v is not None else "" for i, v in enumerate(row)}
                for row in rows[:30] if any(v is not None for v in row)
            ],
            "total_rows": len(rows),
        }
    except Exception as e:
        return {"error": str(e)}

def parse_lead_xls(filepath):
    """Parse Lead .xls file"""
    if xlrd is None:
        try:
            return parse_xlsx_via_pandas(filepath)
        except ImportError:
            return {"error": "xlrd not installed and fallback failed"}

    try:
        book = xlrd.open_workbook(str(filepath))
        sheet = book.sheet_by_index(0)

        rows = []
        for row_idx in range(min(sheet.nrows, 100)):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                row.append(str(cell.value) if cell.value else "")
            rows.append(row)

        return {
            "rows": rows,
            "total_rows": sheet.nrows,
            "total_cols": sheet.ncols,
        }
    except Exception as e:
        return {"error": str(e)}

def parse_xlsx_via_pandas(filepath):
    try:
        import pandas as pd
        df = pd.read_excel(filepath, header=None)
        return {
            "rows": df.head(50).fillna("").astype(str).values.tolist(),
            "total_rows": len(df),
        }
    except ImportError:
        return {"error": "pandas not available"}


print(f"=== WEEK 9 PARSE: 29.06-05.07.2026 ===\n")

# Parse Direct files
print(f"\n--- DIRECT FILES ---")
direct_files = sorted([f for f in os.listdir(DIRECT_DIR) if f.endswith(".xlsx")])
for filename in direct_files:
    filepath = DIRECT_DIR / filename
    account = detect_account(filename)
    print(f"\nFile: {filename}")
    print(f"  Detected account: {account}")

    parsed = parse_xlsx(filepath)
    if parsed is None:
        print("  -> FAILED to parse")
        continue
    if "error" in parsed:
        print(f"  -> ERROR: {parsed['error']}")
        continue

    has_imp = parsed["has_impressions_col"]
    campaigns_count = len(parsed["campaigns"])
    totals = parsed["totals"]
    print(f"  Campaigns: {campaigns_count}")
    print(f"  Has Impressions col: {has_imp}")
    print(f"  Totals: imp={totals['impressions']:.0f}, clicks={totals['clicks']:.0f}, "
          f"spend={totals['spend']:.2f}, conversions={totals['conversions']:.0f}")

    # Save best version per account (prefer with impressions)
    if account:
        if account not in results["direct_accounts"]:
            results["direct_accounts"][account] = {
                "has_impressions_col": has_imp,
                "totals": totals,
                "campaigns": parsed["campaigns"],
                "source_file": filename,
            }
        else:
            # Prefer version with impressions column
            current = results["direct_accounts"][account]
            if has_imp and not current["has_impressions_col"]:
                results["direct_accounts"][account] = {
                    "has_impressions_col": has_imp,
                    "totals": totals,
                    "campaigns": parsed["campaigns"],
                    "source_file": filename,
                }

# Aggregate Direct totals
print(f"\n\n=== AGGREGATED DIRECT TOTALS ===\n")
total_imp = 0
total_clicks = 0
total_spend = 0.0
total_conv = 0

for account in ["e-20010227", "e-17228851", "dune-group", "porg-3uieikjn"]:
    if account in results["direct_accounts"]:
        info = results["direct_accounts"][account]
        t = info["totals"]
        print(f"  {account:20s}: imp={t['impressions']:>7.0f}, clicks={t['clicks']:>5.0f}, "
              f"spend={t['spend']:>10.2f}, conv={t['conversions']:>4.0f}  "
              f"[from: {info['source_file']}]")
        total_imp += t["impressions"]
        total_clicks += t["clicks"]
        total_spend += t["spend"]
        total_conv += t["conversions"]
    else:
        print(f"  {account}: NO DATA")

print(f"\n  {'TOTAL':20s}: imp={total_imp:>7.0f}, clicks={total_clicks:>5.0f}, "
      f"spend={total_spend:>10.2f}, conv={total_conv:>4.0f}")

results["total_impressions"] = total_imp
results["total_clicks"] = total_clicks
results["total_spend"] = total_spend

# Print campaign breakdown for each account
print(f"\n\n=== CAMPAIGN BREAKDOWNS ===\n")
for account in ["e-20010227", "e-17228851", "dune-group", "porg-3uieikjn"]:
    if account not in results["direct_accounts"]:
        continue
    info = results["direct_accounts"][account]
    print(f"\n  --- {account} ---")
    for name, c in sorted(info["campaigns"].items(), key=lambda x: -x[1]["spend"]):
        print(f"    {name[:60]:60s}: imp={c['impressions']:>7.0f}, "
              f"clicks={c['clicks']:>5.0f}, spend={c['spend']:>10.2f}, "
              f"conv={c['conversions']:>4.0f}")

# Parse Webmaster files
print(f"\n\n--- WEBMASTER FILES ---")
for filename in sorted(os.listdir(WEBMASTER_DIR)):
    filepath = WEBMASTER_DIR / filename
    print(f"\nFile: {filename}")
    parsed = parse_seo_webmaster(filepath)
    if "error" in parsed:
        print(f"  -> ERROR: {parsed['error']}")
        continue
    print(f"  Total rows: {parsed['total_rows']}")
    for i, row in enumerate(parsed["rows"][:15]):
        non_empty = {k: v for k, v in row.items() if v}
        if non_empty:
            print(f"    Row {i}: {non_empty}")

# Parse Lead xls
print(f"\n\n--- LEAD XLS ---")
lead_files = sorted([f for f in os.listdir(WEEK9_DIR) if f.endswith(".xls")])
for filename in lead_files:
    filepath = WEEK9_DIR / filename
    print(f"\nFile: {filename}")
    parsed = parse_lead_xls(filepath)
    if "error" in parsed:
        print(f"  -> ERROR: {parsed['error']}")
        continue
    print(f"  Total rows: {parsed.get('total_rows', '?')}")
    rows = parsed.get("rows", [])
    for i, row in enumerate(rows[:30]):
        non_empty = [v for v in row if v]
        if non_empty:
            print(f"    Row {i}: {non_empty[:10]}")

# Save results
out_path = PROJECT_ROOT / "week9_parsed.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"\n\nSaved: {out_path}")
