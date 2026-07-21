"""
Parse W11 data files for 13.07-19.07.2026
Extracts: Direct accounts, SEO data, Lead data
"""
import sys
import io
import os
import json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from pathlib import Path
from openpyxl import load_workbook

WEEK_DIR = Path("0. reports/13-19.07")

ACCOUNT_NAMES = ["e-20010227", "e-17228851", "dune-group", "porg-3uieikjn"]

results = {
    "period": "13.07.2026 - 19.07.2026",
    "direct_accounts": {},
    "seo": {},
    "leads": {},
}

def detect_account(filename):
    for key in ACCOUNT_NAMES:
        if key in filename:
            return key
    return None

def num(v):
    if v is None or v == "" or v == "-":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return 0

print("=" * 70)
print("W11 PARSER - 13.07-19.07.2026")
print("=" * 70)

# ==== 1. PARSE DIRECT FILES ====
print("\n--- DIRECT FILES ---")
all_files = [f for f in os.listdir(WEEK_DIR) if f.endswith(".xlsx") and not f.startswith("dune-group.ru")]
direct_files = []
for f in all_files:
    if detect_account(f):
        direct_files.append(f)

print(f"Direct files: {len(direct_files)}")

for filename in sorted(direct_files):
    filepath = WEEK_DIR / filename
    account = detect_account(filename)
    print(f"\n>>> {filename} | account: {account} <<<")
    if not account:
        continue

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    total_rows = ws.max_row
    print(f"  Total rows: {total_rows}")

    rows_list = list(ws.iter_rows(values_only=True))

    # Find header row
    header_idx = None
    for i, row in enumerate(rows_list[:10]):
        if row and row[0] and "Название кампании" in str(row[0]):
            header_idx = i
            break

    if header_idx is None:
        for i, row in enumerate(rows_list[:20]):
            if row and row[0] and any(kw in str(row[0]) for kw in ["Кампания", "Название", "Campaign"]):
                header_idx = i
                break

    if header_idx is None:
        print(f"  -> No header found.")
        # Dump first rows for debugging
        for i, row in enumerate(rows_list[:5]):
            print(f"    {i}: {[str(v)[:50] if v else '' for v in row]}")
        continue

    headers = list(rows_list[header_idx])
    print(f"  Headers (row {header_idx}): {headers}")

    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

    name_col = col_map.get("Название кампании")
    if name_col is None:
        name_col = 0
    spend_col = col_map.get("Расход, ₽") or col_map.get("Расход")
    clicks_col = col_map.get("Клики")
    impressions_col = col_map.get("Показы")
    conversions_col = col_map.get("Конверсии")

    print(f"  Cols: name={name_col}, spend={spend_col}, clicks={clicks_col}, imp={impressions_col}")

    if name_col is None:
        print(f"  -> No name column, skipping")
        continue

    # Aggregate by campaign
    campaigns = {}
    total_spend = 0.0
    total_clicks = 0
    total_imp = 0
    total_conv = 0

    for row_num, row in enumerate(rows_list[header_idx + 1:], header_idx + 2):
        if not row or name_col >= len(row) or row[name_col] is None:
            continue
        name = str(row[name_col]).strip()
        if name.lower() in ("итого", "total", "", "всего"):
            continue

        spend = num(row[spend_col]) if spend_col is not None and spend_col < len(row) else 0
        clicks = num(row[clicks_col]) if clicks_col is not None and clicks_col < len(row) else 0
        imp = num(row[impressions_col]) if impressions_col is not None and impressions_col < len(row) else 0
        conv = num(row[conversions_col]) if conversions_col is not None and conversions_col < len(row) else 0

        if name in campaigns:
            campaigns[name]["spend"] += spend
            campaigns[name]["clicks"] += clicks
            campaigns[name]["impressions"] += imp
            campaigns[name]["conversions"] += conv
        else:
            campaigns[name] = {"spend": spend, "clicks": clicks, "impressions": imp, "conversions": conv}

        total_spend += spend
        total_clicks += clicks
        total_imp += imp
        total_conv += conv

    campaigns_clean = {k: v for k, v in campaigns.items() if v["clicks"] > 0 or v["spend"] > 0 or v["impressions"] > 0}

    print(f"  Campaigns with data: {len(campaigns_clean)}")
    print(f"  TOTALS: imp={total_imp:.0f}, clicks={total_clicks:.0f}, spend={total_spend:.2f}")

    if campaigns_clean:
        print(f"  Campaigns:")
        for name, c in sorted(campaigns_clean.items(), key=lambda x: -x[1]["spend"]):
            print(f"    {name[:60]:60s} | imp={c['impressions']:>7.0f}, clicks={c['clicks']:>5.0f}, "
                  f"spend={c['spend']:>10.2f}")

    if account not in results["direct_accounts"]:
        results["direct_accounts"][account] = {
            "totals": {"spend": total_spend, "clicks": total_clicks, "impressions": total_imp, "conversions": total_conv},
            "campaigns": campaigns_clean,
            "source_file": filename,
        }

    wb.close()

# ==== 2. PARSE SEO FILES ====
print("\n--- SEO FILES ---")
seo_files = [f for f in os.listdir(WEEK_DIR) if f.startswith("dune-group.ru") and f.endswith(".xlsx")]
print(f"SEO files: {len(seo_files)}")

for filename in seo_files:
    filepath = WEEK_DIR / filename
    print(f"\n>>> {filename} <<<")

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows_list = list(ws.iter_rows(values_only=True))
    print(f"  Total rows: {ws.max_row}")

    # Print first 10 rows to understand structure
    for i, row in enumerate(rows_list[:10]):
        print(f"    {i}: {[str(v)[:40] if v else '' for v in row[:8]]}")

    # Try to find header with visits/clicks
    header_idx = None
    for i, row in enumerate(rows_list[:15]):
        if row and row[0] and any(kw in str(row[0]).lower() for kw in ["страница", "url", "запрос", "query", "источник"]):
            header_idx = i
            break

    if header_idx is None:
        # Just count total rows with data
        data_rows = 0
        for row in rows_list[1:]:
            if any(v is not None and str(v).strip() for v in row[:3]):
                data_rows += 1
        print(f"  Data rows: {data_rows}")
    else:
        headers = [str(h).strip() if h else "" for h in rows_list[header_idx]]
        print(f"  Headers: {headers}")

    wb.close()

# ==== 3. PARSE LEAD FILE ====
print("\n--- LEAD FILE ---")
lead_files = [f for f in os.listdir(WEEK_DIR) if f.startswith("LEAD") and (f.endswith(".xls") or f.endswith(".xlsx"))]
print(f"Lead files: {len(lead_files)}")

for filename in lead_files:
    filepath = WEEK_DIR / filename
    print(f"\n>>> {filename} <<<")

    # Try openpyxl first
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        rows_list = list(ws.iter_rows(values_only=True))
        print(f"  Total rows (openpyxl): {ws.max_row}")

        # Print header row
        for i, row in enumerate(rows_list[:5]):
            print(f"    {i}: {[str(v)[:30] if v else '' for v in row[:12]]}")

        # Find headers
        header_idx = None
        for i, row in enumerate(rows_list[:10]):
            if row and any(str(v).strip() for v in row[:5]):
                # Check if any known header exists
                combined = " ".join(str(v).lower() for v in row if v)
                if any(kw in combined for kw in ["статус", "source", "title", "id", "лид", "lead"]):
                    header_idx = i
                    break

        if header_idx is not None:
            headers = [str(r).strip() if r else "" for r in rows_list[header_idx]]
            print(f"  Headers: {headers}")

            # Count total leads
            leads_count = 0
            target_count = 0
            for row in rows_list[header_idx + 1:]:
                if not row or not any(str(v).strip() for v in row[:3]):
                    continue
                leads_count += 1

            print(f"  Total leads counted: {leads_count}")

        wb.close()
    except Exception as e:
        print(f"  openpyxl failed: {e}")
        # Try xlrd
        try:
            import xlrd
            book = xlrd.open_workbook(str(filepath))
            sheet = book.sheet_by_index(0)
            print(f"  Total rows (xlrd): {sheet.nrows}")
            for i in range(min(10, sheet.nrows)):
                row = [str(sheet.cell(i, j).value)[:30] for j in range(min(12, sheet.ncols))]
                print(f"    {i}: {row}")
        except ImportError:
            print("  xlrd not available")
        except Exception as e2:
            print(f"  xlrd failed: {e2}")

# ==== AGGREGATED TOTALS ====
print("\n" + "=" * 70)
print("AGGREGATED DIRECT TOTALS")
print("=" * 70)
total_imp = total_clicks = total_conversions = 0
total_spend = 0.0
for account in ACCOUNT_NAMES:
    if account in results["direct_accounts"]:
        t = results["direct_accounts"][account]["totals"]
        print(f"  {account:20s}: imp={t['impressions']:>8.0f}, clicks={t['clicks']:>5.0f}, "
              f"spend={t['spend']:>11.2f}")
        total_imp += t["impressions"]
        total_clicks += t["clicks"]
        total_spend += t["spend"]
        total_conversions += t["conversions"]
    else:
        print(f"  {account}: NO DATA")

print(f"  {'TOTAL':20s}: imp={total_imp:>8.0f}, clicks={total_clicks:>5.0f}, "
      f"spend={total_spend:>11.2f}")

results["total_impressions"] = total_imp
results["total_clicks"] = total_clicks
results["total_spend"] = total_spend
results["total_conversions"] = total_conversions

# Save
out_path = Path("week11_parsed.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)
print(f"\nSaved: {out_path}")
