"""Debug: inspect actual structure of W9 files"""
import os
from pathlib import Path
from openpyxl import load_workbook

DIRECT_DIR = Path("05.07/111")
WEBMASTER_DIR = Path("05.07/webmaster")

# Inspect Direct xlsx files
print("=" * 70)
print("DIRECT FILES - Raw inspection")
print("=" * 70)

direct_files = sorted([f for f in os.listdir(DIRECT_DIR) if f.endswith(".xlsx")
                        and "Поисковые" not in f])
for filename in direct_files[:4]:
    filepath = DIRECT_DIR / filename
    print(f"\n>>> {filename} <<<")
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:15]):
        # Filter empty
        non_empty = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
        if non_empty:
            truncated = [(j, str(v)[:80]) for j, v in non_empty]
            print(f"  Row {i}: {truncated}")
    wb.close()

# Inspect Webmaster xlsx files
print("\n" + "=" * 70)
print("WEBMASTER FILES")
print("=" * 70)
for filename in sorted(os.listdir(WEBMASTER_DIR)):
    filepath = WEBMASTER_DIR / filename
    print(f"\n>>> {filename} <<<")
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:20]):
        non_empty = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
        if non_empty:
            truncated = [(j, str(v)[:80]) for j, v in non_empty]
            print(f"  Row {i}: {truncated}")
    wb.close()

# Inspect "Поисковые системы" file
print("\n" + "=" * 70)
print("ПОISKOVYE SISTEMY (Master Search) FILE")
print("=" * 70)
ps_file = DIRECT_DIR / "Поисковые системы-2026-06-29-2026-07-05.xlsx"
if ps_file.exists():
    wb = load_workbook(ps_file, data_only=True, read_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:20]):
        non_empty = [(j, v) for j, v in enumerate(row) if v is not None and str(v).strip()]
        if non_empty:
            truncated = [(j, str(v)[:80]) for j, v in non_empty]
            print(f"  Row {i}: {truncated}")
    wb.close()

# Inspect the .xls file as bytes
print("\n" + "=" * 70)
print("LEAD XLS - raw bytes")
print("=" * 70)
with open("05.07/LEAD_20260706_6454f36c_6a4bdebd6ed11.xls", "rb") as f:
    head = f.read(400)
    print(f"First 400 bytes: {head}")
    print(f"Length: {f.tell()}")
