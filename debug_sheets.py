"""Deep debug - W9 direct xlsx files - show all sheets and rows"""
import os
from pathlib import Path
from openpyxl import load_workbook

DIRECT_DIR = Path("05.07/111")

print("=" * 70)
print("LIST ALL SHEETS in each W9 Direct file")
print("=" * 70)

for filename in sorted(os.listdir(DIRECT_DIR)):
    if not filename.endswith(".xlsx") or "Поисковые" in filename:
        continue
    filepath = DIRECT_DIR / filename
    print(f"\n>>> {filename} <<<")
    wb = load_workbook(filepath, data_only=True, read_only=False)
    print(f"  Sheet names: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  Sheet '{sheet_name}': max_row={ws.max_row}, max_col={ws.max_column}")
        # Read first 8 rows
        if ws.max_row > 0:
            count = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 8 or i >= ws.max_row - 5:
                    non_empty = [(j, str(v)[:80]) for j, v in enumerate(row) if v is not None and str(v).strip()]
                    if non_empty:
                        print(f"    Row {i}: {non_empty[:6]}")
                count += 1
            print(f"  Iterated {count} rows")
    wb.close()
