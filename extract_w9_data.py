"""Extract actual data from W9 files - SEO + Leads"""
import os
from pathlib import Path
from openpyxl import load_workbook

DIRECT_DIR = Path("05.07/111")

# === 1. Parse Поисковые системы file (SEO) ===
print("=" * 70)
print("ПОИСКОВЫЕ СИСТЕМЫ (SEO/WEBMASTER DATA)")
print("=" * 70)

ps_file = DIRECT_DIR / "Поисковые системы-2026-06-29-2026-07-05.xlsx"
wb = load_workbook(ps_file, data_only=True, read_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

seo_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if any(v is not None for v in row):
        row_strs = [str(v)[:60] if v is not None else "" for v in row]
        print(f"  Row {i:2d}: {row_strs}")
        seo_rows.append(row_strs)
wb.close()

# Aggregate SEO visits
seo_total_visits = 0
print("\n--- SEO SUMMATION ---")
for row in seo_rows[5:20]:  # skip header
    if len(row) >= 3 and row[2]:
        try:
            visits = float(str(row[2]).replace(",", "."))
            seo_total_visits += visits
            print(f"  {row[0][:30] if len(row) > 0 else ''}: visits={visits}")
        except (ValueError, TypeError):
            pass
print(f"\n  SEO TOTAL VISITS: {seo_total_visits}")

# === 2. Parse HTML Lead file ===
print("\n" + "=" * 70)
print("LEAD XLS (actually HTML)")
print("=" * 70)

import re
from html.parser import HTMLParser

with open("05.07/LEAD_20260706_6454f36c_6a4bdebd6ed11.xls", "r", encoding="utf-8") as f:
    html_content = f.read()

print(f"Total HTML length: {len(html_content)}")

# Quick parse: extract table rows
class TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_table = False
        self.in_row = False
        self.in_cell = False
        self.current_cells = []
        self.current_row = []
        self.all_rows = []
        self.headers = []
        self.in_head = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self.in_table = True
        elif tag == "thead":
            self.in_head = True
        elif tag == "tr":
            self.in_row = True
            self.current_row = []
        elif tag in ("td", "th"):
            self.in_cell = True
            self.current_cells.append([])

    def handle_endtag(self, tag):
        if tag == "table":
            self.in_table = False
        elif tag == "thead":
            self.in_head = False
        elif tag == "tr":
            if self.in_head and not self.headers:
                # First thead row = headers
                self.headers = [" ".join(c).strip() for c in self.current_row]
            self.all_rows.append(list(self.current_row))
            self.in_row = False
        elif tag in ("td", "th"):
            self.in_cell = False

    def handle_data(self, data):
        if self.in_cell and self.current_cells:
            self.current_cells[-1].append(data.strip())
        elif self.in_row:
            self.current_row.append(" ".join(self.current_cells[-1]) if self.current_cells else "")

parser = TableParser()
parser.feed(html_content)

print(f"Headers: {parser.headers}")
print(f"Total rows: {len(parser.all_rows)}")

# Show first 30 leads
print("\n--- FIRST 30 LEADS ---")
for i, row in enumerate(parser.all_rows[:30]):
    cells = [" ".join(c).strip() if isinstance(c, list) else str(c).strip() for c in row]
    print(f"  Row {i}: {cells}")

# Try to count total - usually WAY more than what display showed
print(f"\n--- TAIL (last 5 rows) ---")
for i, row in enumerate(parser.all_rows[-5:]):
    cells = [" ".join(c).strip() if isinstance(c, list) else str(c).strip() for c in row]
    print(f"  Row {len(parser.all_rows)-5+i}: {cells}")

# Now classify: find leading-name column, stage column, date column, source column
# Typical lead columns: Название лида | Стадия | Дата создания | Источник | ID
print("\n--- COLUMN INDEXING ---")
print(f"Headers: {parser.headers}")
if parser.headers:
    for i, h in enumerate(parser.headers):
        print(f"  Col {i}: {h}")
