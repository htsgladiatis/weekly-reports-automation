"""Fixed W9 parser - properly reads all rows of large xlsx"""
import sys
import io
import os
import json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from pathlib import Path
from openpyxl import load_workbook

DIRECT_DIR = Path("05.07/111")
WEBMASTER_DIR = Path("05.07/webmaster")

ACCOUNT_NAMES = ["e-20010227", "e-17228851", "dune-group", "porg-3uieikjn"]

results = {
    "period": "29.06.2026 - 05.07.2026",
    "direct_accounts": {},
}

def detect_account(filename):
    for key in ACCOUNT_NAMES:
        if key in filename:
            return key
    return None

def num(v):
    if v is None or v == "" or v == "-":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return 0

print("=" * 70)
print("FIXED W9 PARSER - reads all rows properly")
print("=" * 70)

# Parse Direct files
direct_files = sorted([f for f in os.listdir(DIRECT_DIR) if f.endswith(".xlsx") and "Поисковые" not in f])
print(f"\nDirect files found: {len(direct_files)}")

for filename in direct_files:
    filepath = DIRECT_DIR / filename
    account = detect_account(filename)
    print(f"\n>>> {filename} | account: {account} <<<")
    if not account:
        continue

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    total_rows = ws.max_row
    print(f"  Total rows: {total_rows}")

    # Read header row (typically row 4 with column names)
    # Find headers
    header_idx = None
    rows_list = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows_list[:10]):
        if row and row[0] and "Название кампании" in str(row[0]):
            header_idx = i
            break

    if header_idx is None:
        # Try alternate patterns
        for i, row in enumerate(rows_list[:20]):
            if row and row[0] and any(kw in str(row[0]) for kw in ["Кампания", "Название", "Campaign"]):
                header_idx = i
                break

    if header_idx is None:
        print(f"  -> No header found, dumping rows 0..10:")
        for i, row in enumerate(rows_list[:10]):
            print(f"    {i}: {[str(v)[:50] if v else '' for v in row]}")
        continue

    headers = list(rows_list[header_idx])
    print(f"  Headers (row {header_idx}): {headers}")

    # Column map
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

    name_col = col_map.get("Название кампании")
    if name_col is None:
        # Fallback: take column 0 as name
        for i, h in enumerate(headers):
            if h is not None:
                name_col = i
                break
    spend_col = col_map.get("Расход, ₽") or col_map.get("Расход")
    clicks_col = col_map.get("Клики")
    impressions_col = col_map.get("Показы")
    conversions_col = col_map.get("Конверсии")

    print(f"  Cols: name={name_col}, spend={spend_col}, clicks={clicks_col}, imp={impressions_col}, conv={conversions_col}")

    if name_col is None:
        print(f"  -> CRITICAL: no name column, skipping")
        continue

    # Aggregate
    campaigns = {}
    total_spend = 0.0
    total_clicks = 0
    total_imp = 0
    total_conv = 0

    for row in rows_list[header_idx + 1:]:
        if not row or name_col >= len(row) or row[name_col] is None:
            continue
        name = str(row[name_col]).strip()
        # Skip itogo rows
        if name.lower() in ("итого", "total", ""):
            continue

        spend = num(row[spend_col]) if spend_col is not None and spend_col < len(row) else 0
        clicks = num(row[clicks_col]) if clicks_col is not None and clicks_col < len(row) else 0
        imp = num(row[impressions_col]) if impressions_col is not None and impressions_col < len(row) else 0
        conv = num(row[conversions_col]) if conversions_col is not None and conversions_col < len(row) else 0

        if name in campaigns:
            campaigns[name]["spend"] += spend
            campaigns[name]["clicks"] += clicks
            campaigns[name]["impressions"] += imp
            campaigns[name]["conversions"] += conv
        else:
            campaigns[name] = {"spend": spend, "clicks": clicks, "impressions": imp, "conversions": conv}

        total_spend += spend
        total_clicks += clicks
        total_imp += imp
        total_conv += conv

    # Cleanup: ignore empty campaigns and aggregate totals
    campaigns_clean = {k: v for k, v in campaigns.items() if v["clicks"] > 0 or v["spend"] > 0 or v["impressions"] > 0}

    print(f"\n  Total Campaigns: {len(campaigns_clean)}")
    print(f"  TOTALS: imp={total_imp:.0f}, clicks={total_clicks:.0f}, spend={total_spend:.2f}, conv={total_conv:.0f}")

    print(f"\n  Campaign breakdown (sorted by spend):")
    for name, c in sorted(campaigns_clean.items(), key=lambda x: -x[1]["spend"]):
        print(f"    {name[:60]:60s} | imp={c['impressions']:>7.0f}, clicks={c['clicks']:>5.0f}, "
              f"spend={c['spend']:>10.2f}, conv={c['conversions']:>4.0f}")

    # Save best version
    if account not in results["direct_accounts"]:
        results["direct_accounts"][account] = {
            "totals": {"spend": total_spend, "clicks": total_clicks, "impressions": total_imp, "conversions": total_conv},
            "campaigns": campaigns_clean,
            "source_file": filename,
            "has_impressions_col": impressions_col is not None,
        }

    wb.close()

# Aggregate totals
print("\n" + "=" * 70)
print("AGGREGATED DIRECT TOTALS")
print("=" * 70)
total_imp = total_clicks = total_conversions = 0
total_spend = 0.0
for account in ACCOUNT_NAMES:
    if account in results["direct_accounts"]:
        t = results["direct_accounts"][account]["totals"]
        src = results["direct_accounts"][account]["source_file"]
        print(f"  {account:20s}: imp={t['impressions']:>8.0f}, clicks={t['clicks']:>5.0f}, "
              f"spend={t['spend']:>11.2f}, conv={t['conversions']:>4.0f}  [{src}]")
        total_imp += t["impressions"]
        total_clicks += t["clicks"]
        total_spend += t["spend"]
        total_conversions += t["conversions"]
    else:
        print(f"  {account}: NO DATA")

print(f"  {'TOTAL':20s}: imp={total_imp:>8.0f}, clicks={total_clicks:>5.0f}, "
      f"spend={total_spend:>11.2f}, conv={total_conversions:>4.0f}")

results["total_impressions"] = total_imp
results["total_clicks"] = total_clicks
results["total_spend"] = total_spend
results["total_conversions"] = total_conversions

# Save
out_path = Path("week9_parsed.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)
print(f"\nSaved: {out_path}")
