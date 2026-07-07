import openpyxl
import os
import glob
from collections import defaultdict

folder = "1-7.06"
out = open("xlsx_aggregated.txt", "w", encoding="utf-8")

def process_file(fpath, account_name):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    
    # Find header row
    header_row = None
    header_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row[0] and 'Название кампании' in str(row[0]):
            header_row = list(row)
            header_idx = i
            break
    
    if not header_row:
        out.write(f"  ERROR: header not found\n")
        wb.close()
        return
    
    out.write(f"  Headers: {header_row}\n\n")
    
    # Column indices
    cols = {str(v): i for i, v in enumerate(header_row) if v is not None}
    out.write(f"  Column map: {cols}\n\n")
    
    # Aggregate by campaign
    campaigns = defaultdict(lambda: {'imp': 0.0, 'clicks': 0.0, 'spend': 0.0, 'conv': 0.0})
    
    for row in ws.iter_rows(min_row=header_idx+1, values_only=True):
        name = row[0]
        if not name or name == 'Итого':
            continue
        
        def safe_float(val):
            if val is None or val == '-':
                return 0.0
            try:
                return float(val)
            except:
                return 0.0
        
        imp = safe_float(row[cols.get('Показы')] if 'Показы' in cols else None)
        clicks = safe_float(row[cols.get('Клики')] if 'Клики' in cols else None)
        spend = safe_float(row[cols.get('Расход, ₽')] if 'Расход, ₽' in cols else None)
        conv = safe_float(row[cols.get('Конверсии')] if 'Конверсии' in cols else None)
        
        campaigns[name]['imp'] += imp
        campaigns[name]['clicks'] += clicks
        campaigns[name]['spend'] += spend
        campaigns[name]['conv'] += conv
    
    # Print results
    total_imp = total_clicks = total_spend = total_conv = 0
    out.write(f"  {'Кампания':<70} {'Показы':>10} {'Клики':>8} {'Расход':>10} {'Конв':>6}\n")
    out.write(f"  {'-'*110}\n")
    for name, d in sorted(campaigns.items()):
        out.write(f"  {name:<70} {d['imp']:>10.0f} {d['clicks']:>8.0f} {d['spend']:>10.2f} {d['conv']:>6.0f}\n")
        total_imp += d['imp']
        total_clicks += d['clicks']
        total_spend += d['spend']
        total_conv += d['conv']
    out.write(f"  {'ИТОГО':<70} {total_imp:>10.0f} {total_clicks:>8.0f} {total_spend:>10.2f} {total_conv:>6.0f}\n")
    wb.close()

# Process each "Новый" file (they have daily breakdown + impressions)
novyi_files = {
    'e-20010227': None,
    'e-17228851': None,
    'dune-group': None,
    'porg-3uieikjn': None,
}

all_files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
for fpath in all_files:
    fname = os.path.basename(fpath)
    for acc in novyi_files:
        if acc in fname and 'Новый' in fname:
            novyi_files[acc] = fpath

out.write("=== AGGREGATED DATA FROM 'Новый' FILES (with impressions) ===\n\n")
for acc, fpath in novyi_files.items():
    out.write(f"\n{'='*60}\n")
    out.write(f"ACCOUNT: {acc}\n")
    out.write(f"FILE: {os.path.basename(fpath) if fpath else 'NOT FOUND'}\n")
    out.write(f"{'='*60}\n")
    if fpath:
        process_file(fpath, acc)

# Also check simple files for e-17228851 (no "Новый" version with impressions might be missing)
out.write("\n\n=== SIMPLE FILES (no impressions column) ===\n\n")
simple_map = {}
for fpath in all_files:
    fname = os.path.basename(fpath)
    for acc in ['e-20010227', 'e-17228851', 'dune-group', 'porg-3uieikjn']:
        if acc in fname and 'Новый' not in fname:
            simple_map[acc] = fpath

for acc, fpath in simple_map.items():
    out.write(f"\n{'='*60}\n")
    out.write(f"SIMPLE - ACCOUNT: {acc}\n")
    out.write(f"FILE: {os.path.basename(fpath)}\n")
    out.write(f"{'='*60}\n")
    process_file(fpath, acc)

out.close()
print("Done, see xlsx_aggregated.txt")
